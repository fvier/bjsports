from html import escape
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def brl(value):
    formatted = f'{float(value):,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
    return f'R$ {formatted}'


def build_financial_markdown(report, detailed=False):
    lines = [
        '# Relatório de Mensalidades — BJ Sports',
        '',
        f'**Período:** {report["period_label"]}',
        f'**Gerado em:** {report["generated_at"]}',
        f'**Responsável pela emissão:** {report["generated_by"]} (@{report["generated_by_username"]})',
        f'**Perfil de acesso:** {report["generated_by_role"]}',
        f'**Tipo do relatório:** {"Detalhado" if detailed else "Resumido"}',
        f'**Escopo:** {report["scope_label"]}',
        '',
        '## Indicadores consolidados',
        '',
        '| Indicador | Valor |',
        '|---|---:|',
        f'| Alunos incluídos | {report["student_count"]} |',
        f'| Alunos isentos | {report["exempt_count"]} |',
        f'| Valor previsto | {brl(report["totals"]["expected"])} |',
        f'| Valor recebido | {brl(report["totals"]["received"])} |',
        f'| Taxa de recebimento | {report["receipt_rate"]:.1f}% |',
        f'| Valor em atraso | {brl(report["totals"]["overdue"])} |',
        f'| Valor a vencer | {brl(report["totals"]["future"])} |',
        f'| Sem registro financeiro | {brl(report["totals"]["unrecorded"])} |',
        '',
        '## Resumo por competência',
        '',
        '| Competência | Alunos | Isentos | Previsto | Recebido | Em atraso | A vencer | Sem registro |',
        '|---|---:|---:|---:|---:|---:|---:|---:|',
    ]
    for row in report['summary_rows']:
        lines.append(
            f'| {row["period"]} | {row["students"]} | {row["exempt"]} | {brl(row["expected"])} | '
            f'{brl(row["received"])} | {brl(row["overdue"])} | '
            f'{brl(row["future"])} | {brl(row["unrecorded"])} |'
        )

    if detailed:
        lines.extend([
            '',
            '## Lançamentos detalhados',
            '',
            '| Competência | Aluno | Usuário | Plano | Vencimento | Status | Valor | Baixa |',
            '|---|---|---|---|---:|---|---:|---|',
        ])
        for row in report['detail_rows']:
            lines.append(
                f'| {row["period"]} | {row["student"]} | @{row["username"]} | '
                f'{row["plan"]} | Dia {row["due_date"]} | {row["status_label"]} | '
                f'{brl(row["amount"])} | {row["paid_at"] or "—"} |'
            )

    lines.extend([
        '',
        '---',
        'Documento administrativo gerado pelo BJ Sports ERP.',
    ])
    return '\n'.join(lines)


def _table_cells(lines):
    rows = []
    for line in lines:
        cells = [cell.strip() for cell in line.strip().strip('|').split('|')]
        if all(cell and set(cell) <= {'-', ':'} for cell in cells):
            continue
        rows.append(cells)
    return rows


def _markdown_table(lines, body_style, header_style, available_width):
    raw_rows = _table_cells(lines)
    rows = [
        [Paragraph(escape(cell), header_style if row_index == 0 else body_style) for cell in row]
        for row_index, row in enumerate(raw_rows)
    ]
    column_count = max((len(row) for row in rows), default=1)
    if column_count == 7:
        proportions = [.14, .08, .156, .156, .156, .156, .156]
    elif column_count == 8 and raw_rows and 'Isentos' in raw_rows[0]:
        proportions = [.12, .07, .07, .148, .148, .148, .148, .148]
    elif column_count == 8:
        proportions = [.105, .19, .115, .205, .09, .115, .105, .075]
    else:
        proportions = [1 / column_count] * column_count
    widths = [available_width * proportion for proportion in proportions]
    table = Table(rows, colWidths=widths, repeatRows=1, hAlign='LEFT')
    commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#172033')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#B42318')),
        ('GRID', (0, 0), (-1, -1), .3, colors.HexColor('#D8DEE8')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    for row_index in range(1, len(rows)):
        commands.append(('BACKGROUND', (0, row_index), (-1, row_index),
                         colors.HexColor('#F8FAFC' if row_index % 2 else '#FFFFFF')))
    table.setStyle(TableStyle(commands))
    return table


def _indicator_cards(lines, label_style, value_style, available_width):
    entries = _table_cells(lines)[1:]
    accent_colors = {
        'Alunos incluídos': '#3B82F6',
        'Alunos isentos': '#7C3AED',
        'Valor previsto': '#64748B',
        'Valor recebido': '#16A34A',
        'Taxa de recebimento': '#CA8A04',
        'Valor em atraso': '#DC2626',
        'Valor a vencer': '#2563EB',
    }
    column_count = 4 if len(entries) == 8 else 3
    card_width = (available_width - (4 * column_count)) / column_count
    cards = []
    for label, value in entries:
        accent = colors.HexColor(accent_colors.get(label, '#64748B'))
        card = Table([
            [Paragraph(escape(label.upper()), label_style)],
            [Paragraph(escape(value), value_style)],
        ], colWidths=[card_width], rowHeights=[18, 28])
        card.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
            ('BOX', (0, 0), (-1, -1), .5, colors.HexColor('#D8DEE8')),
            ('LINEABOVE', (0, 0), (-1, 0), 3, accent),
            ('LEFTPADDING', (0, 0), (-1, -1), 9),
            ('RIGHTPADDING', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        cards.append(card)
    while len(cards) % column_count:
        cards.append('')
    grid = Table([cards[index:index + column_count] for index in range(0, len(cards), column_count)],
                 colWidths=[card_width] * column_count, hAlign='LEFT')
    grid.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    return grid


def _brand_header(title, title_style, subtitle_style, available_width):
    logo_path = Path(__file__).resolve().parent / 'static' / 'img' / 'logo_inverted.png'
    logo = Image(str(logo_path), width=18 * mm, height=22.5 * mm) if logo_path.exists() else ''
    copy = [
        Paragraph('BJ SPORTS · CENTRO DE TREINAMENTO', subtitle_style),
        Paragraph(escape(title), title_style),
        Paragraph('RELATÓRIO FINANCEIRO E CONTROLE DE MENSALIDADES', subtitle_style),
    ]
    header = Table([[logo, copy]], colWidths=[25 * mm, available_width - (25 * mm)])
    header.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#111827')),
        ('LINEBELOW', (0, 0), (-1, -1), 3, colors.HexColor('#DC2626')),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return header


def _plain_markdown(line):
    return line.replace('**', '').strip()


def markdown_to_pdf(markdown_text):
    output = BytesIO()
    page_size = landscape(A4)
    margin = 13 * mm
    lines = markdown_text.splitlines()
    responsible_line = next((_plain_markdown(line).split(':', 1)[-1].strip()
                             for line in lines if line.startswith('**Responsável pela emissão:**')), 'Não informado')
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=15 * mm,
        bottomMargin=16 * mm,
        title='Relatório de Mensalidades — BJ Sports',
        author=responsible_line,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='ReportTitle', parent=styles['Title'], fontName='Helvetica-Bold',
                              fontSize=16, leading=19, textColor=colors.white, alignment=0, spaceAfter=2))
    styles.add(ParagraphStyle(name='ReportBrand', parent=styles['BodyText'], fontName='Helvetica-Bold',
                              fontSize=6.5, leading=8, textColor=colors.HexColor('#CBD5E1'), spaceAfter=2))
    styles.add(ParagraphStyle(name='ReportH2', parent=styles['Heading2'], fontName='Helvetica-Bold',
                              fontSize=11, leading=14, textColor=colors.HexColor('#172033'), spaceBefore=9, spaceAfter=5))
    styles.add(ParagraphStyle(name='ReportBody', parent=styles['BodyText'], fontName='Helvetica',
                              fontSize=7.4, leading=9.5, textColor=colors.HexColor('#475467')))
    styles.add(ParagraphStyle(name='ReportTable', parent=styles['BodyText'], fontName='Helvetica',
                              fontSize=6.1, leading=7.4, textColor=colors.HexColor('#344054')))
    styles.add(ParagraphStyle(name='ReportTableHead', parent=styles['BodyText'], fontName='Helvetica-Bold',
                              fontSize=6.2, leading=7.4, textColor=colors.white, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='CardLabel', parent=styles['BodyText'], fontName='Helvetica-Bold',
                              fontSize=5.8, leading=7, textColor=colors.HexColor('#667085')))
    styles.add(ParagraphStyle(name='CardValue', parent=styles['BodyText'], fontName='Helvetica-Bold',
                              fontSize=12, leading=14, textColor=colors.HexColor('#172033')))

    story = []
    index = 0
    available_width = page_size[0] - (2 * margin)
    current_section = ''
    while index < len(lines):
        line = lines[index].strip()
        if not line:
            story.append(Spacer(1, 3))
        elif line.startswith('|'):
            table_lines = []
            while index < len(lines) and lines[index].strip().startswith('|'):
                table_lines.append(lines[index].strip())
                index += 1
            if current_section == 'Indicadores consolidados':
                story.append(_indicator_cards(table_lines, styles['CardLabel'], styles['CardValue'], available_width))
            else:
                story.append(_markdown_table(table_lines, styles['ReportTable'], styles['ReportTableHead'], available_width))
            story.append(Spacer(1, 7))
            continue
        elif line.startswith('# '):
            story.append(_brand_header(line[2:], styles['ReportTitle'], styles['ReportBrand'], available_width))
            story.append(Spacer(1, 8))
        elif line.startswith('## '):
            current_section = line[3:]
            story.append(Paragraph(escape(current_section), styles['ReportH2']))
        elif line == '---':
            story.append(Spacer(1, 5))
        else:
            safe = escape(line).replace('**', '<b>', 1).replace('**', '</b>', 1)
            story.append(Paragraph(safe, styles['ReportBody']))
        index += 1

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#D8DEE8'))
        canvas.setLineWidth(.4)
        canvas.line(margin, 11 * mm, page_size[0] - margin, 11 * mm)
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#667085'))
        canvas.drawString(margin, 7 * mm, f'BJ Sports ERP · Emitido por {responsible_line}')
        canvas.drawRightString(page_size[0] - margin, 7 * mm, f'Página {doc.page}')
        canvas.restoreState()

    def continued_page(canvas, doc):
        footer(canvas, doc)
        canvas.saveState()
        canvas.setFont('Helvetica-Bold', 7)
        canvas.setFillColor(colors.HexColor('#172033'))
        canvas.drawString(margin, page_size[1] - (9 * mm), 'BJ SPORTS · RELATÓRIO FINANCEIRO')
        canvas.setStrokeColor(colors.HexColor('#DC2626'))
        canvas.setLineWidth(1)
        canvas.line(margin, page_size[1] - (11 * mm), page_size[0] - margin, page_size[1] - (11 * mm))
        canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=continued_page)
    output.seek(0)
    return output


def financial_report_to_xlsx(report, detailed=False):
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Resumo'
    summary.append(['Relatório de Mensalidades — BJ Sports'])
    summary.append(['Período', report['period_label']])
    summary.append(['Gerado em', report['generated_at']])
    summary.append(['Responsável', f'{report["generated_by"]} (@{report["generated_by_username"]})'])
    summary.append(['Perfil de acesso', report['generated_by_role']])
    summary.append(['Escopo', report['scope_label']])
    summary.append([])
    summary.append(['Competência', 'Alunos', 'Isentos', 'Previsto', 'Recebido', 'Em atraso', 'A vencer', 'Sem registro'])
    for row in report['summary_rows']:
        summary.append([row['period'], row['students'], row['exempt'], row['expected'], row['received'],
                        row['overdue'], row['future'], row['unrecorded']])
    summary.append([])
    summary.append(['TOTAL', report['student_count'], report['exempt_count'], report['totals']['expected'], report['totals']['received'],
                    report['totals']['overdue'], report['totals']['future'], report['totals']['unrecorded']])

    if detailed:
        details = workbook.create_sheet('Detalhado')
        headers = ['Competência', 'Aluno', 'Usuário', 'Contato', 'Plano', 'Vencimento', 'Status', 'Valor', 'Baixa']
        details.append(headers)
        for row in report['detail_rows']:
            details.append([row['period'], row['student'], row['username'], row['contact'], row['plan'],
                            f'Dia {row["due_date"]}', row['status_label'], row['amount'], row['paid_at'] or ''])
        _style_sheet(details, header_row=1, currency_columns=(8,))

    _style_sheet(summary, header_row=8, currency_columns=(4, 5, 6, 7, 8))
    summary['A1'].font = Font(size=16, bold=True, color='172033')
    summary.merge_cells('A1:H1')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _style_sheet(sheet, header_row, currency_columns=()):
    header_fill = PatternFill('solid', fgColor='172033')
    for cell in sheet[header_row]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.freeze_panes = f'A{header_row + 1}'
    sheet.auto_filter.ref = f'A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}'
    for column in currency_columns:
        for row in range(header_row + 1, sheet.max_row + 1):
            sheet.cell(row, column).number_format = 'R$ #,##0.00'
    for index, column_cells in enumerate(sheet.columns, 1):
        width = min(36, max(12, max(len(str(cell.value or '')) for cell in column_cells) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
