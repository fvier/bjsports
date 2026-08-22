import os
import unittest
from io import BytesIO
from datetime import datetime, timedelta
from unittest.mock import patch
from openpyxl import load_workbook

os.environ['DATABASE_URL'] = 'sqlite:///:memory:'
os.environ['SECRET_KEY'] = 'test-only-secret-key'
os.environ.pop('SEED_DEMO_DATA', None)

from app import (app, db, User, Plan, MonthlyPayment, Attendance, Booking, ClassGroup, ClassEnrollment,
                 SpecialClassEvent, PushSubscription, GraduationRecord, InternalChampionship,
                 ChampionshipRegistration, ChampionshipWeightDivision,
                 ChampionshipWeightRevision, ChampionshipMatch, ChampionshipScoreEvent,
                 ContractAcceptance, MEMBERSHIP_TERMS_VERSION, PRIVACY_NOTICE_VERSION,
                 calendar_token_serializer)


class BJSportsTestCase(unittest.TestCase):
    def setUp(self):
        app.config.update(TESTING=True)
        self.client = app.test_client()
        with app.app_context():
            db.drop_all()
            db.create_all()
            db.session.add(Plan(name='Plano Teste', category='Planos Individuais', price='R$ 100,00/mês'))
            for index, (username, role) in enumerate([('aluno', 'aluno'), ('monitor', 'monitor'), ('instrutor', 'instrutor')], 1):
                user = User(username=username, name=username.title(), cpf=f'000.000.00{index}-00',
                            ddd='83', phone='999999999', plan='Plano Teste — R$ 100,00/mês',
                            due_date='5', start_month=1, role=role, payment_status='Em Dia')
                user.set_password('senha-segura')
                db.session.add(user)
            db.session.commit()

    def csrf(self):
        self.client.get('/login', follow_redirects=True)
        with self.client.session_transaction() as data:
            return data['_csrf_token']

    def login(self, username, password='senha-segura'):
        return self.client.post('/login', data={'action': 'login', 'portalCpf': username,
            'portalPassword': password, 'csrf_token': self.csrf()})

    def test_invalid_login_does_not_create_session(self):
        self.assertEqual(self.login('inexistente', 'errada').status_code, 302)
        with self.client.session_transaction() as data:
            self.assertNotIn('user_id', data)

    def test_login_redirects_to_welcome_dashboard(self):
        response = self.login('aluno')
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers['Location'].endswith('/dashboard'))
        dashboard = self.client.get('/dashboard')
        self.assertEqual(dashboard.status_code, 200)
        self.assertIn('Resumo da sua conta', dashboard.get_data(as_text=True))
        self.assertIn('Matriculado há menos de 1 mês', dashboard.get_data(as_text=True))
        self.assertIn('data-flash-duration="2800"', dashboard.get_data(as_text=True))
        self.assertIn('data-flash-close', dashboard.get_data(as_text=True))

    def test_registration_requires_versioned_terms_and_image_authorization(self):
        page = self.client.get('/login.html?mode=register').get_data(as_text=True)
        self.assertIn('TERMO DE ADESÃO', page)
        self.assertIn('Seu contrato, sem letras miúdas', page)
        self.assertIn('Cancelamento da matrícula', page)
        self.assertIn('Saúde física, saúde mental e bem-estar', page)
        self.assertIn('Seus direitos sobre os dados', page)
        self.assertIn('Alunos menores de idade', page)
        self.assertIn('<strong>Enquanto existir débito vencido, o sistema poderá impedir novos check-ins e reservas.</strong>', page)
        self.assertIn('<strong>A BJ Sports definiu esta autorização como uma condição para concluir uma nova matrícula por este cadastro.</strong>', page)
        self.assertIn('Imagem de crianças e adolescentes', page)
        self.assertIn('melhor interesse do menor', page)
        self.assertIn('name="acceptMembershipTerms"', page)
        self.assertIn('name="acknowledgePrivacy"', page)
        self.assertIn('name="confirmLegalCapacity"', page)
        self.assertIn('name="imageConsentScope"', page)
        self.assertIn('value="minor_guardian"', page)
        self.assertIn('Revise e confirme suas escolhas', page)
        self.assertIn('Nenhuma foto será enviada agora.', page)
        self.assertNotIn('Prefiro não aparecer', page)
        self.assertIn('Autorizo como aluno adulto', page)
        self.assertIn('Autorizo pelo aluno menor', page)
        self.assertIn('Sem a autorização, o cadastro não será concluído.', page)
        self.assertIn('name="regEmail"', page)
        self.assertIn('name="regSex"', page)
        self.assertIn('Prefiro não informar', page)
        self.assertIn('<option value="" selected disabled>Selecione uma modalidade</option>', page)
        self.assertNotIn('Plano Passe Livre', page)
        self.assertIn('data-plan-schedule="ter-qui"', page)
        self.assertIn('data-plan-schedule="seg-qua-sex"', page)
        self.assertIn('data-plan-schedule="todos"', page)
        self.assertIn('Todos dias', page)
        self.assertLess(page.index('Seg - Qua - Sex'), page.index('Ter &amp; Qui'))
        self.assertIn('data-schedule-price', page)
        self.assertIn('const selectedPrices = selectedPlan?.prices || {};', page)
        self.assertIn('renderPlanOptions();', page)
        self.assertIn('Cadastro de aluno menor de idade', page)
        self.assertNotIn('👶', page)
        self.assertIn('Aula Particular', page)
        self.assertEqual(page.count('name="comboModalities"'), 3)
        self.assertIn('id="comboModalityFields"', page)
        self.assertIn('name="privateInstructor"', page)
        self.assertIn('<optgroup label="Professores">', page)
        self.assertIn('<optgroup label="Monitores">', page)
        self.assertLess(page.index('<optgroup label="Professores">'), page.index('<optgroup label="Monitores">'))
        self.assertIn('class="is-approved"', page)
        with app.app_context():
            plan = Plan.query.one()
            plan_value = f'{plan.name} — {plan.price}'

        registration = {
            'action': 'register', 'regUsername': 'novoaluno', 'regName': 'Novo Aluno',
            'regCpf': '529.982.247-25', 'regDDD': '83', 'regPhoneNumber': '999999999',
            'regEmail': 'novoaluno@example.com', 'regSex': 'prefer_not',
            'regBirthDate': '1990-05-10',
            'regPlan': plan_value, 'regTrainingDays': 'seg-qua-sex',
            'regDueDate': '15', 'regPass': 'senha-segura',
        }
        denied = self.client.post('/login', data={
            **registration, 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('Leia e aceite o termo de adesão', denied.get_data(as_text=True))
        self.assertIn('mode=register', denied.request.url)
        with app.app_context():
            self.assertIsNone(User.query.filter_by(username='novoaluno').first())

        denied_image = self.client.post('/login', data={
            **registration, 'acceptMembershipTerms': 'on', 'acknowledgePrivacy': 'on',
            'confirmLegalCapacity': 'on', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('autorize o uso de imagem', denied_image.get_data(as_text=True))
        with app.app_context():
            self.assertIsNone(User.query.filter_by(username='novoaluno').first())

        accepted = self.client.post('/login', data={
            **registration, 'acceptMembershipTerms': 'on', 'acknowledgePrivacy': 'on',
            'confirmLegalCapacity': 'on', 'imageConsentScope': 'adult',
            'csrf_token': self.csrf(),
        })
        self.assertEqual(accepted.status_code, 302)
        self.assertTrue(accepted.headers['Location'].endswith('/dashboard'))
        with app.app_context():
            user = User.query.filter_by(username='novoaluno').one()
            self.assertEqual(user.membership_terms_version, MEMBERSHIP_TERMS_VERSION)
            self.assertIsNotNone(user.membership_terms_accepted_at)
            self.assertEqual(user.privacy_notice_version, '2026-08-17')
            self.assertIsNotNone(user.privacy_notice_accepted_at)
            self.assertTrue(user.image_use_consent)
            self.assertIsNotNone(user.image_use_consent_at)
            self.assertEqual(user.image_consent_scope, 'adult')
            self.assertEqual(user.email, 'novoaluno@example.com')
            self.assertEqual(user.sex, 'prefer_not')
            self.assertEqual(user.plan, 'Plano Teste • Seg, Qua, Sex — R$ 100,00/mês')
            acceptance = ContractAcceptance.query.filter_by(user_id=user.id).one()
            self.assertEqual(acceptance.source, 'registration')
            self.assertEqual(acceptance.membership_terms_version, MEMBERSHIP_TERMS_VERSION)

    def test_minor_image_consent_requires_and_records_legal_guardian(self):
        with app.app_context():
            plan = Plan.query.one()
            plan_value = f'{plan.name} — {plan.price}'
        registration = {
            'action': 'register', 'regUsername': 'alunomenor', 'regName': 'Aluno Menor',
            'regCpf': '111.444.777-35', 'regDDD': '83', 'regPhoneNumber': '988888888',
            'regEmail': 'responsavel@example.com', 'regSex': 'feminino',
            'regBirthDate': '2012-05-10',
            'regPlan': plan_value, 'regTrainingDays': 'ter-qui',
            'regDueDate': '5', 'regPass': 'senha-segura',
            'acceptMembershipTerms': 'on', 'acknowledgePrivacy': 'on',
            'confirmLegalCapacity': 'on', 'imageConsentScope': 'minor_guardian',
        }
        denied = self.client.post('/login', data={
            **registration, 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('nome completo do responsável', denied)
        with app.app_context():
            self.assertIsNone(User.query.filter_by(username='alunomenor').first())

        accepted = self.client.post('/login', data={
            **registration, 'imageGuardianName': 'Maria Responsável',
            'imageGuardianCpf': '529.982.247-25', 'imageGuardianRelationship': 'mae',
            'csrf_token': self.csrf(),
        })
        self.assertEqual(accepted.status_code, 302)
        with app.app_context():
            user = User.query.filter_by(username='alunomenor').one()
            self.assertTrue(user.image_use_consent)
            self.assertEqual(user.image_consent_scope, 'minor_guardian')
            self.assertEqual(user.image_consent_guardian_name, 'Maria Responsável')
            self.assertEqual(user.image_consent_guardian_cpf, '529.982.247-25')
            self.assertEqual(user.image_consent_guardian_relationship, 'mae')
            self.assertIsNotNone(user.image_use_consent_at)

    def test_private_class_requires_and_records_selected_professional(self):
        payload = {
            'action': 'register', 'regUsername': 'particular', 'regName': 'Aluno Particular',
            'regCpf': '529.982.247-25', 'regDDD': '83', 'regPhoneNumber': '988887777',
            'regEmail': 'particular@example.com', 'regSex': 'prefer_not',
            'regBirthDate': '1990-05-10',
            'regPlan': '__private_class__', 'regTrainingDays': 'ter-qui',
            'regDueDate': '15', 'regPass': 'senha-segura', 'acceptMembershipTerms': 'on',
            'acknowledgePrivacy': 'on', 'confirmLegalCapacity': 'on',
            'imageConsentScope': 'adult', 'csrf_token': self.csrf(),
        }
        denied = self.client.post('/login', data=payload, follow_redirects=True)
        self.assertIn('Escolha um professor ou monitor válido', denied.get_data(as_text=True))

        payload['privateInstructor'] = 'instrutor'
        payload['csrf_token'] = self.csrf()
        accepted = self.client.post('/login', data=payload)
        self.assertEqual(accepted.status_code, 302)
        with app.app_context():
            user = User.query.filter_by(username='particular').one()
            self.assertEqual(user.plan, 'Aula Particular com Instrutor • Ter, Qui')

    def test_combo_plus_two_requires_three_distinct_modalities(self):
        with app.app_context():
            combo = Plan(name='Plano Multitreino Premium', category='Combos & Planos Especiais',
                         price='R$ 180,00/mês', selection_count=3, force_all_days=True,
                         modality='Jiu-Jitsu, Boxe, Muay Thai, MMA', sub='Plano configurável',
                         features='Treinos integrados;Acompanhamento técnico')
            db.session.add(combo)
            db.session.commit()
            combo_value = f'{combo.name} — {combo.price}'
        payload = {
            'action': 'register', 'regUsername': 'alunocombo', 'regName': 'Aluno Combo',
            'regCpf': '111.444.777-35', 'regDDD': '83', 'regPhoneNumber': '988887777',
            'regEmail': 'combo@example.com', 'regSex': 'prefer_not', 'regBirthDate': '1990-05-10',
            'regPlan': combo_value, 'regTrainingDays': 'todos', 'regDueDate': '15',
            'regPass': 'senha-segura', 'acceptMembershipTerms': 'on', 'acknowledgePrivacy': 'on',
            'confirmLegalCapacity': 'on', 'imageConsentScope': 'adult',
        }
        denied = self.client.post('/login', data={**payload, 'comboModalities': ['Jiu-Jitsu', 'Jiu-Jitsu'], 'csrf_token': self.csrf()}, follow_redirects=True)
        self.assertIn('Escolha 3 modalidades diferentes', denied.get_data(as_text=True))

        registration_page = self.client.get('/login?mode=register').get_data(as_text=True)
        self.assertIn('Plano Multitreino Premium', registration_page)
        self.assertIn('"selection_count": 3', registration_page)
        self.assertIn('"features": [', registration_page)
        self.assertIn('registrationPlanDetails', registration_page)

        accepted = self.client.post('/login', data={**payload, 'comboModalities': ['Jiu-Jitsu', 'Boxe', 'MMA'], 'csrf_token': self.csrf()})
        self.assertEqual(accepted.status_code, 302)
        with app.app_context():
            user = User.query.filter_by(username='alunocombo').one()
            self.assertEqual(user.get_selected_modalities_list(), ['Jiu-Jitsu', 'Boxe', 'MMA'])
            self.assertEqual(user.birth_date.isoformat(), '1990-05-10')
            combo_user_id = user.id

        forbidden = self.client.get('/gestao/modalidades-combo', follow_redirects=True)
        self.assertIn('Somente professores e monitores', forbidden.get_data(as_text=True))

        self.login('monitor')
        management = self.client.get('/gestao/modalidades-combo')
        self.assertEqual(management.status_code, 200)
        self.assertIn('Aluno Combo', management.get_data(as_text=True))
        updated = self.client.post('/gestao/modalidades-combo', data={
            'user_id': combo_user_id, 'combo_modalities': ['Muay Thai', 'MMA', 'Boxe'],
            'csrf_token': self.csrf(),
        })
        self.assertEqual(updated.status_code, 302)
        with app.app_context():
            user = db.session.get(User, combo_user_id)
            self.assertEqual(user.get_selected_modalities_list(), ['Muay Thai', 'MMA', 'Boxe'])

    def test_classes_page_is_public(self):
        response = self.client.get('/turmas')
        self.assertEqual(response.status_code, 200)
        page = response.get_data(as_text=True)
        self.assertIn('classes-page', page)
        self.assertIn('Jiu-Jitsu Kids 2', page)
        self.assertNotIn('Jiu-Jitsu Kids 2 (a partir de 8 anos)', page)
        self.assertIn('class="experimental-checkbox-label"', page)
        self.assertIn('<strong>Aula Experimental Grátis</strong>', page)
        self.assertIn('experimental-checkbox-action', page)

    def test_weekly_calendar_is_available_to_logged_users_and_supports_filters(self):
        self.assertEqual(self.client.get('/calendario').status_code, 302)
        self.login('aluno')
        response = self.client.get('/calendario')
        self.assertEqual(response.status_code, 200)
        page = response.get_data(as_text=True)
        self.assertIn('Calendário de Aulas', page)
        self.assertIn('<strong>34</strong> aula(s) nesta semana', page)
        self.assertIn('Boxe Matinal', page)
        self.assertIn('Muay Thai Kids', page)
        self.assertEqual(page.count('Sem aulas'), 2)
        self.assertIn('Horários livres', page)
        self.assertIn('06:00–21:30', page)
        self.assertNotIn('Adicionar aula especial', page)
        self.assertNotIn('Gerenciar turmas', page)

        filtered = self.client.get('/calendario?modality=Boxe').get_data(as_text=True)
        self.assertIn('Boxe Matinal', filtered)
        self.assertNotIn('Jiu-Jitsu Tarde', filtered)

    def test_monitor_can_add_special_event_without_schedule_conflicts(self):
        self.login('monitor')
        page = self.client.get('/calendario').get_data(as_text=True)
        self.assertIn('Adicionar aula especial', page)
        self.assertNotIn('Gerenciar turmas', page)

        today = datetime.now().date()
        free_saturday = today + timedelta(days=(5 - today.weekday()) % 7)
        if free_saturday < today:
            free_saturday += timedelta(days=7)
        response = self.client.post('/calendario', data={
            'title': 'Treino aberto de sábado', 'event_date': free_saturday.isoformat(),
            'start_time': '10:00', 'end_time': '11:00', 'event_modality': 'Treino especial',
            'notes': 'Evento de teste', 'week': free_saturday.isoformat(),
            'modality': 'todas', 'audience': 'todos', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        rendered = response.get_data(as_text=True)
        self.assertIn('Aula especial adicionada ao calendário com sucesso!', rendered)
        self.assertIn('Treino aberto de sábado', rendered)
        self.assertIn('10:00–11:00', rendered)
        with app.app_context():
            self.assertEqual(SpecialClassEvent.query.count(), 1)

        next_monday = today + timedelta(days=(7 - today.weekday()) % 7)
        if next_monday < today:
            next_monday += timedelta(days=7)
        blocked = self.client.post('/calendario', data={
            'title': 'Treino conflitante', 'event_date': next_monday.isoformat(),
            'start_time': '06:30', 'end_time': '07:15', 'event_modality': 'Boxe',
            'week': next_monday.isoformat(), 'modality': 'todas', 'audience': 'todos',
            'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Conflito de horário com Boxe Matinal (06:00–07:00)', blocked)
        with app.app_context():
            self.assertEqual(SpecialClassEvent.query.count(), 1)

        self.client.post('/logout', data={'csrf_token': self.csrf()})
        self.login('aluno')
        denied = self.client.post('/calendario', data={
            'title': 'Evento sem permissão', 'event_date': free_saturday.isoformat(),
            'start_time': '12:00', 'end_time': '13:00', 'event_modality': 'Treino especial',
            'week': free_saturday.isoformat(), 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Somente monitores e instrutores podem adicionar aulas especiais.', denied)
        with app.app_context():
            self.assertEqual(SpecialClassEvent.query.count(), 1)

    def test_personal_calendar_and_push_only_use_logged_users_enrollments(self):
        self.login('aluno')
        page = self.client.get('/calendario').get_data(as_text=True)
        self.assertIn('Lembretes e Google', page)
        self.assertIn('Somente as turmas das quais você participa serão usadas.', page)
        self.assertIn('Baixar .ICS', page)

        with app.app_context():
            user = User.query.filter_by(username='aluno').first()
            enrolled_names = [item.class_group.name for item in user.class_enrollments if item.active]
            self.assertEqual(len(enrolled_names), 1)
            token = calendar_token_serializer.dumps({'user_id': user.id})
        calendar_response = self.client.get(f'/calendario/minhas-turmas.ics?token={token}&download=1')
        calendar_text = calendar_response.get_data(as_text=True)
        self.assertEqual(calendar_response.status_code, 200)
        self.assertIn('text/calendar', calendar_response.content_type)
        self.assertIn('BEGIN:VALARM', calendar_text)
        self.assertIn('RRULE:FREQ=WEEKLY', calendar_text)
        self.assertIn(enrolled_names[0], calendar_text)
        self.assertTrue(all(name not in calendar_text for name in ['Boxe Matinal', 'Muay Thai Kids']
                            if name != enrolled_names[0]))
        self.assertEqual(self.client.get('/calendario/minhas-turmas.ics?token=invalido').status_code, 404)

        with patch.dict(os.environ, {'VAPID_PUBLIC_KEY': 'test-public-key'}):
            enabled = self.client.post('/api/calendario/push', json={
                'endpoint': 'https://push.example/subscription-1',
                'keys': {'p256dh': 'client-public-key', 'auth': 'client-auth'},
            }, headers={'X-CSRF-Token': self.csrf()})
            self.assertEqual(enabled.status_code, 200)
            with app.app_context():
                subscription = PushSubscription.query.one()
                self.assertEqual(subscription.user.username, 'aluno')
                self.assertTrue(subscription.active)
            disabled = self.client.delete('/api/calendario/push', json={},
                                          headers={'X-CSRF-Token': self.csrf()})
            self.assertEqual(disabled.status_code, 200)
            with app.app_context():
                self.assertFalse(PushSubscription.query.one().active)

    def test_monitor_creates_championship_and_student_registers_from_card(self):
        self.login('monitor')
        today = datetime.now().date()
        event_date = today + timedelta(days=30)
        deadline = today + timedelta(days=20)
        created = self.client.post('/campeonatos/interno', data={
            'name': 'Interno BJ Sports Teste', 'modality': 'Jiu-Jitsu',
            'event_date': event_date.isoformat(), 'registration_deadline': deadline.isoformat(),
            'location': 'Sede BJ Sports', 'max_participants': '16',
            'description': 'Evento interno para validação.', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        created_page = created.get_data(as_text=True)
        self.assertIn('criado com sucesso!', created_page)
        self.assertIn('Interno BJ Sports Teste', created_page)
        self.assertIn('Ver e fazer inscrição', created_page)
        with app.app_context():
            championship = InternalChampionship.query.one()
            championship_id = championship.id
            self.assertEqual(championship.created_by_username, 'monitor')

        self.client.post('/logout', data={'csrf_token': self.csrf()})
        self.login('aluno')
        student_page = self.client.get('/campeonatos/interno').get_data(as_text=True)
        self.assertIn(f'/campeonatos/interno/{championship_id}', student_page)
        self.assertNotIn('Novo campeonato', student_page)
        registered = self.client.post(f'/campeonatos/interno/{championship_id}', data={
            'weight': '72.5', 'age_division': 'Adulto', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        registered_page = registered.get_data(as_text=True)
        self.assertIn('Inscrição confirmada', registered_page)
        self.assertIn('72.5 kg', registered_page)
        with app.app_context():
            registration = ChampionshipRegistration.query.one()
            self.assertEqual(registration.user.username, 'aluno')
            self.assertEqual(registration.championship_id, championship_id)
            self.assertEqual(registration.belt_color_snapshot, 'branca')

        duplicate = self.client.post(f'/campeonatos/interno/{championship_id}', data={
            'weight': '73', 'age_division': 'Adulto', 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Você já está inscrito neste campeonato.', duplicate)
        with app.app_context():
            self.assertEqual(ChampionshipRegistration.query.count(), 1)

    def test_weight_tables_are_visible_and_editable_by_monitor_with_history(self):
        self.login('aluno')
        student_response = self.client.get('/campeonatos/pesos')
        student_page = student_response.get_data(as_text=True)
        self.assertEqual(student_response.status_code, 200)
        self.assertIn('MASCULINO — FBJJMMA', student_page)
        self.assertIn('FEMININO — FBJJMMA', student_page)
        self.assertIn('Até 57,5 kg', student_page)
        self.assertIn('role="tablist"', student_page)
        self.assertIn('data-weight-tab="masculino"', student_page)
        self.assertIn('data-weight-tab="feminino"', student_page)
        self.assertIn('data-weight-panel="feminino" hidden', student_page)
        self.assertNotIn('Edição liberada', student_page)
        with app.app_context():
            self.assertEqual(ChampionshipWeightDivision.query.count(), 18)
            division = ChampionshipWeightDivision.query.filter_by(gender='masculino', category='Galo').one()
            division_id = division.id

        denied = self.client.post('/campeonatos/pesos', data={
            'division_id': division_id, 'category': 'Galo', 'pre_mirim': '', 'mirim': '',
            'infantil': 'Até 32,2 kg', 'infanto': 'Até 44,3 kg', 'juvenil': 'Até 53,5 kg',
            'adulto_master': 'Até 58 kg', 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Somente monitores e instrutores podem editar', denied)

        self.client.post('/logout', data={'csrf_token': self.csrf()})
        self.login('monitor')
        updated = self.client.post('/campeonatos/pesos', data={
            'division_id': division_id, 'category': 'Galo', 'pre_mirim': '', 'mirim': '',
            'infantil': 'Até 32,2 kg', 'infanto': 'Até 44,3 kg', 'juvenil': 'Até 53,5 kg',
            'adulto_master': 'Até 58 kg', 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Categoria Galo da tabela masculino atualizada.', updated)
        self.assertIn('Até 58 kg', updated)
        with app.app_context():
            division = db.session.get(ChampionshipWeightDivision, division_id)
            revision = ChampionshipWeightRevision.query.one()
            self.assertEqual(division.adulto_master, 'Até 58 kg')
            self.assertEqual(revision.updated_by_username, 'monitor')
            self.assertIn('Até 57,5 kg', revision.changes_json)

    def test_monitor_controls_scoreboard_while_student_has_read_only_access(self):
        with app.app_context():
            championship = InternalChampionship(
                name='Interno do Tatame', modality='Jiu-Jitsu',
                event_date=datetime.now().date() + timedelta(days=15),
                registration_deadline=datetime.now().date() + timedelta(days=10),
                location='Sede BJ Sports', max_participants=24,
                created_by_username='monitor',
            )
            db.session.add(championship)
            db.session.commit()
            championship_id = championship.id

        self.login('monitor')
        free_timer = self.client.get('/campeonatos/placar-e-timer.html?view=timer').get_data(as_text=True)
        self.assertIn('data-standalone-clock', free_timer)
        self.assertIn('data-timer-fullscreen', free_timer)
        self.assertIn('data-standalone-action="toggle"', free_timer)
        self.assertIn('data-standalone-custom', free_timer)
        self.assertIn('Tempo personalizado', free_timer)
        self.assertIn('Sons: Início • 30s • Fim', free_timer)
        self.assertIn('Modo livre: funciona imediatamente no navegador', free_timer)
        for minutes in (2, 3, 4, 5, 6, 10):
            self.assertIn(f'data-standalone-duration="{minutes}"', free_timer)

        created = self.client.post('/campeonatos/placar', data={
            'action': 'create_match', 'championship_id': championship_id,
            'red_competitor': 'João Silva', 'blue_competitor': 'Carlos Souza',
            'category': 'Adulto • Faixa Azul • Leve', 'mat_area': 'Área 1',
            'duration_minutes': '5', 'penalty_limit': '4', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        monitor_page = created.get_data(as_text=True)
        self.assertEqual(created.status_code, 200)
        self.assertIn('pronto para o placar.', monitor_page)
        self.assertIn('João Silva', monitor_page)
        self.assertIn('Carlos Souza', monitor_page)
        self.assertIn('05:00', monitor_page)
        self.assertIn('Nova Luta', monitor_page)
        self.assertIn('Montada ou costas', monitor_page)
        self.assertIn('Passagem de guarda', monitor_page)
        self.assertIn('Queda, raspagem ou joelho na barriga', monitor_page)
        self.assertIn('CORNER VERMELHO', monitor_page)
        self.assertIn('CORNER AZUL', monitor_page)
        self.assertIn('data-scoreboard-stage-fullscreen', monitor_page)
        self.assertIn('TEMPO DA LUTA', monitor_page)
        self.assertIn('Iniciar luta', monitor_page)
        self.assertIn('value="red_disqualify"', monitor_page)
        self.assertIn('value="blue_disqualify"', monitor_page)
        self.assertIn('ÁREAS E CONFRONTOS', monitor_page)
        self.assertIn('Padrão — desclassifica na 4ª', monitor_page)
        self.assertIn('Sub-15 — desclassifica na 6ª', monitor_page)
        self.assertNotIn('value="red_add_2"', monitor_page)
        self.assertIn('data-scoreboard-tab="placar"', monitor_page)
        self.assertIn('data-scoreboard-tab="timer"', monitor_page)
        self.assertIn('>Cronômetro</strong>', monitor_page)
        self.assertIn('id="scoreboardPanelTimer"', monitor_page)
        for minutes in (2, 3, 4, 5, 6, 10):
            self.assertIn(f'value="set_duration_{minutes}"', monitor_page)
        self.assertNotIn('value="set_duration_8"', monitor_page)
        with app.app_context():
            match_id = ChampionshipMatch.query.one().id

        changed_duration = self.client.post('/campeonatos/placar', data={
            'action': 'set_duration_10', 'match_id': match_id, 'scoreboard_view': 'timer',
            'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('view=timer', changed_duration.request.url)
        changed_duration_page = changed_duration.get_data(as_text=True)
        self.assertIn('10:00', changed_duration_page)
        self.assertIn('data-scoreboard-tab="timer" aria-controls="scoreboardPanelTimer" aria-selected="true"', changed_duration_page)
        self.assertIn('CRONÔMETRO OFICIAL', changed_duration_page)
        with app.app_context():
            match = db.session.get(ChampionshipMatch, match_id)
            self.assertEqual(match.duration_seconds, 600)
            self.assertEqual(match.remaining_seconds, 600)
            self.assertFalse(match.timer_running)

        custom_duration = self.client.post('/campeonatos/placar', data={
            'action': 'set_custom_duration', 'match_id': match_id,
            'custom_minutes': '1', 'custom_seconds': '30', 'scoreboard_view': 'timer',
            'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('01:30', custom_duration.get_data(as_text=True))
        with app.app_context():
            match = db.session.get(ChampionshipMatch, match_id)
            self.assertEqual((match.duration_seconds, match.remaining_seconds), (90, 90))

        invalid_duration = self.client.post('/campeonatos/placar', data={
            'action': 'set_custom_duration', 'match_id': match_id,
            'custom_minutes': '0', 'custom_seconds': '0', 'scoreboard_view': 'timer',
            'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Informe um tempo válido entre 00:01 e 99:59.', invalid_duration)

        for action in ('red_two_points', 'red_advantage', 'blue_penalty', 'start_timer',
                       'pause_timer', 'finish_match'):
            response = self.client.post('/campeonatos/placar', data={
                'action': action, 'match_id': match_id, 'csrf_token': self.csrf(),
            })
            self.assertEqual(response.status_code, 302)
        with app.app_context():
            match = db.session.get(ChampionshipMatch, match_id)
            self.assertEqual(match.red_score, 2)
            self.assertEqual(match.red_advantages, 1)
            self.assertEqual(match.blue_penalties, 1)
            self.assertEqual(match.status, 'finalizada')
            self.assertEqual(match.winner_side, 'red')
            self.assertEqual(match.updated_by_username, 'monitor')
            self.assertEqual(ChampionshipScoreEvent.query.filter_by(match_id=match_id).count(), 3)

        self.client.post('/logout', data={'csrf_token': self.csrf()})
        self.login('aluno')
        student_page = self.client.get(f'/campeonatos/placar?match={match_id}').get_data(as_text=True)
        self.assertIn('João Silva', student_page)
        self.assertIn('Placar e Timer', student_page)
        self.assertIn('Vencedor', student_page)
        self.assertNotIn('Nova luta', student_page)
        denied = self.client.post('/campeonatos/placar', data={
            'action': 'reopen_match', 'match_id': match_id, 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Somente monitores e instrutores podem controlar o placar.', denied)
        with app.app_context():
            self.assertEqual(db.session.get(ChampionshipMatch, match_id).status, 'finalizada')

    def test_scoreboard_penalties_apply_consequences_and_support_undo(self):
        with app.app_context():
            championship = InternalChampionship(
                name='Teste de Punições', modality='Jiu-Jitsu',
                event_date=datetime.now().date() + timedelta(days=5),
                registration_deadline=datetime.now().date() + timedelta(days=2),
                location='Sede', max_participants=10, created_by_username='monitor',
            )
            db.session.add(championship)
            db.session.flush()
            match = ChampionshipMatch(
                championship_id=championship.id, red_competitor='Vermelho',
                blue_competitor='Azul', category='Adulto', mat_area='Área 1',
            )
            db.session.add(match)
            db.session.commit()
            match_id = match.id

        self.login('monitor')
        for expected_penalties in range(1, 5):
            response = self.client.post('/campeonatos/placar', data={
                'action': 'red_penalty', 'match_id': match_id,
                'scoreboard_view': 'placar', 'csrf_token': self.csrf(),
            })
            self.assertEqual(response.status_code, 302)
            with app.app_context():
                match = db.session.get(ChampionshipMatch, match_id)
                self.assertEqual(match.red_penalties, expected_penalties)
                if expected_penalties == 2:
                    self.assertEqual(match.blue_advantages, 1)
                if expected_penalties == 3:
                    self.assertEqual(match.blue_score, 2)
        with app.app_context():
            match = db.session.get(ChampionshipMatch, match_id)
            self.assertEqual((match.status, match.winner_side), ('finalizada', 'blue'))
            self.assertEqual(ChampionshipScoreEvent.query.filter_by(match_id=match_id).count(), 4)

        undone = self.client.post('/campeonatos/placar', data={
            'action': 'undo_score_event', 'match_id': match_id,
            'scoreboard_view': 'placar', 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Lançamento desfeito: Punição.', undone)
        with app.app_context():
            match = db.session.get(ChampionshipMatch, match_id)
            self.assertEqual(match.red_penalties, 3)
            self.assertNotEqual(match.status, 'finalizada')
            self.assertIsNone(match.winner_side)
            self.assertEqual(ChampionshipScoreEvent.query.filter_by(match_id=match_id, undone_at=None).count(), 3)

        disqualified = self.client.post('/campeonatos/placar', data={
            'action': 'red_disqualify', 'match_id': match_id,
            'scoreboard_view': 'placar', 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Desclassificação direta', disqualified)
        with app.app_context():
            match = db.session.get(ChampionshipMatch, match_id)
            self.assertEqual((match.status, match.winner_side), ('finalizada', 'blue'))

    def test_graduation_page_and_navigation_are_removed(self):
        self.login('aluno')
        self.assertEqual(self.client.get('/graduacao').status_code, 404)
        self.assertEqual(self.client.get('/graduacao.html').status_code, 404)
        dashboard = self.client.get('/dashboard').get_data(as_text=True)
        self.assertNotIn('Graduação & Faixas', dashboard)
        self.assertNotIn('Minha graduação', dashboard)

    def test_instructor_manages_graduation_with_auditable_history(self):
        self.login('aluno')
        self.assertEqual(self.client.get('/gestao/graduacoes').status_code, 302)
        with app.app_context():
            student_id = User.query.filter_by(username='aluno').first().id
        with self.client.session_transaction() as data:
            data.clear()
        self.login('instrutor')
        page = self.client.get('/gestao/graduacoes')
        self.assertEqual(page.status_code, 200)
        self.assertIn('Administração de Graduação', page.get_data(as_text=True))
        response = self.client.post('/gestao/graduacoes', data={
            'user_id': student_id, 'belt_color': 'azul', 'belt_degree': '1',
            'graduation_date': datetime.now().date().isoformat(),
            'notes': 'Avaliação técnica concluída', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        rendered = response.get_data(as_text=True)
        self.assertIn('atualizada para faixa Azul, 1º grau', rendered)
        self.assertIn('Avaliação técnica concluída', rendered)
        with app.app_context():
            student = db.session.get(User, student_id)
            record = GraduationRecord.query.one()
            self.assertEqual((student.belt_color, student.belt_degree), ('azul', 1))
            self.assertEqual((record.previous_belt_color, record.previous_belt_degree), ('branca', 0))
            self.assertEqual((record.new_belt_color, record.new_belt_degree), ('azul', 1))
            self.assertEqual(record.updated_by_username, 'instrutor')

    def test_anonymous_and_student_cannot_access_admin(self):
        self.assertEqual(self.client.get('/mensalidades_admin').status_code, 302)
        self.assertEqual(self.client.get('/integracoes/catraca').status_code, 302)
        self.assertEqual(self.client.get('/gestao/graduacoes').status_code, 302)
        self.assertEqual(self.client.get('/campeonatos/interno').status_code, 302)
        self.login('aluno')
        self.assertEqual(self.client.get('/mensalidades_admin').status_code, 302)
        self.assertEqual(self.client.get('/gestao').status_code, 302)
        self.assertEqual(self.client.get('/integracoes/catraca').status_code, 302)
        self.assertEqual(self.client.get('/campeonatos/interno').status_code, 200)
        self.assertEqual(self.client.get('/campeonatos/pesos').status_code, 200)
        self.assertEqual(self.client.get('/campeonatos/placar').status_code, 200)
        current_period = datetime.now().strftime('%Y-%m')
        self.assertEqual(self.client.post('/relatorios/mensalidades/exportar', data={
            'format': 'pdf', 'report_type': 'summary', 'scope': 'all',
            'period_start': current_period, 'period_end': current_period,
            'csrf_token': self.csrf(),
        }).status_code, 302)

    def test_roles_can_access_authorized_pages(self):
        self.login('monitor')
        self.assertEqual(self.client.get('/presencas').status_code, 200)
        self.assertEqual(self.client.get('/mensalidades_admin').status_code, 302)
        self.assertEqual(self.client.get('/financeiro_dashboard').status_code, 302)
        self.assertEqual(self.client.get('/integracoes/pagamentos').status_code, 302)
        self.assertEqual(self.client.get('/integracoes/catraca').status_code, 302)
        self.assertEqual(self.client.get('/campeonatos/interno').status_code, 200)
        self.assertEqual(self.client.get('/campeonatos/pesos').status_code, 200)
        self.assertEqual(self.client.get('/campeonatos/placar').status_code, 200)
        self.assertEqual(self.client.get('/gestao/turmas').status_code, 302)
        self.assertEqual(self.client.get('/planos_admin').status_code, 302)
        self.client.post('/logout', data={'csrf_token': self.csrf()})
        self.login('instrutor')
        self.assertEqual(self.client.get('/mensalidades_admin').status_code, 200)
        self.assertEqual(self.client.get('/financeiro_dashboard').status_code, 200)
        self.assertEqual(self.client.get('/integracoes/pagamentos').status_code, 200)
        self.assertEqual(self.client.get('/integracoes/catraca').status_code, 200)
        self.assertEqual(self.client.get('/gestao/graduacoes').status_code, 200)
        self.assertEqual(self.client.get('/campeonatos/interno').status_code, 200)
        self.assertEqual(self.client.get('/campeonatos/placar').status_code, 200)
        self.assertEqual(self.client.get('/gestao/turmas').status_code, 200)
        self.assertEqual(self.client.get('/planos_admin').status_code, 200)
        self.assertEqual(self.client.get('/gestao').status_code, 200)

    def test_financial_reports_export_pdf_via_markdown_and_detailed_xlsx(self):
        self.login('instrutor')
        current_period = datetime.now().strftime('%Y-%m')
        base_data = {
            'period_start': current_period,
            'period_end': current_period,
            'scope': 'all',
            'csrf_token': self.csrf(),
        }

        pdf_response = self.client.post('/relatorios/mensalidades/exportar', data={
            **base_data, 'format': 'pdf', 'report_type': 'summary'
        })
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response.mimetype, 'application/pdf')
        self.assertEqual(pdf_response.headers['X-Report-Workflow'], 'Markdown-to-PDF')
        self.assertTrue(pdf_response.data.startswith(b'%PDF'))
        self.assertIn(b'/Subtype /Image', pdf_response.data)
        self.assertIn('.pdf', pdf_response.headers['Content-Disposition'])

        xlsx_response = self.client.post('/relatorios/mensalidades/exportar', data={
            **base_data, 'format': 'xlsx', 'report_type': 'detailed'
        })
        self.assertEqual(xlsx_response.status_code, 200)
        self.assertTrue(xlsx_response.data.startswith(b'PK'))
        workbook = load_workbook(BytesIO(xlsx_response.data), read_only=True)
        self.assertEqual(workbook.sheetnames, ['Resumo', 'Detalhado'])
        self.assertEqual(workbook['Resumo']['A1'].value, 'Relatório de Mensalidades — BJ Sports')
        self.assertEqual(workbook['Resumo']['B4'].value, 'Instrutor (@instrutor)')
        self.assertEqual(workbook['Detalhado']['A1'].value, 'Competência')

    def test_financial_report_rejects_invalid_period(self):
        self.login('instrutor')
        response = self.client.post('/relatorios/mensalidades/exportar', data={
            'format': 'pdf', 'report_type': 'summary',
            'period_start': '2026-08', 'period_end': '2026-07',
            'scope': 'all', 'csrf_token': self.csrf(),
        })
        self.assertEqual(response.status_code, 400)
        self.assertIn('período final', response.get_json()['error'])

    def test_instructor_can_toggle_monthly_fee_exemption_without_counting_values(self):
        self.login('instrutor')
        current_date = datetime.now()
        current_period = current_date.strftime('%Y-%m')
        previous_date = current_date.replace(day=1) - timedelta(days=1)
        previous_period = previous_date.strftime('%Y-%m')
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            student.payment_status = 'Pendente'
            student.set_month_status(str(previous_date.month), 'pago', previous_date.year)
            student.set_month_status(str(current_date.month), 'atrasado', current_date.year)
            db.session.commit()
            student_id = student.id

        response = self.client.post('/mensalidades_admin', data={
            'action': 'toggle_exemption', 'user_id': student_id,
            'is_exempt': '1', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        page = response.get_data(as_text=True)
        self.assertIn('agora está isento(a) de mensalidade', page)
        self.assertIn('ISENTO ATUAL (R$ 0,00)', page)
        self.assertIn('fee-exemption-form', page)
        with app.app_context():
            student = db.session.get(User, student_id)
            self.assertTrue(student.monthly_fee_exempt)
            self.assertEqual(student.fee_exempted_by_username, 'instrutor')
            self.assertEqual(student.get_numeric_price(), 0.0)
            self.assertEqual(student.get_overdue_details()['total_debt'], 0.0)
            self.assertFalse(student.has_overdue_payments())
            paid_payment = MonthlyPayment.query.filter_by(
                user_id=student.id, year=previous_date.year, month=previous_date.month
            ).one()
            self.assertEqual(paid_payment.status, 'pago')
            self.assertEqual(float(paid_payment.amount), 100.0)

        report = self.client.post('/relatorios/mensalidades/exportar', data={
            'format': 'xlsx', 'report_type': 'detailed', 'scope': 'all',
            'period_start': previous_period, 'period_end': current_period,
            'csrf_token': self.csrf(),
        })
        workbook = load_workbook(BytesIO(report.data), read_only=True)
        rows = list(workbook['Detalhado'].iter_rows(values_only=True))
        paid_row = next(row for row in rows[1:] if row[0] == previous_date.strftime('%m/%Y') and row[2] == 'aluno')
        exempt_row = next(row for row in rows[1:] if row[0] == current_date.strftime('%m/%Y') and row[2] == 'aluno')
        self.assertEqual(paid_row[6], 'Pago')
        self.assertEqual(paid_row[7], 100)
        self.assertEqual(exempt_row[6], 'Isento')
        self.assertEqual(exempt_row[7], 0)

        with self.client.session_transaction() as data:
            data.clear()
        self.login('aluno')
        student_page = self.client.get('/mensalidades_aluno').get_data(as_text=True)
        self.assertIn('ALUNO ISENTO DE MENSALIDADE', student_page)
        dashboard = self.client.get('/dashboard').get_data(as_text=True)
        self.assertIn('Sem novas cobranças mensais', dashboard)

        with self.client.session_transaction() as data:
            data.clear()
        self.login('instrutor')
        self.client.post('/mensalidades_admin', data={
            'action': 'toggle_exemption', 'user_id': student_id,
            'is_exempt': '0', 'csrf_token': self.csrf(),
        })
        with app.app_context():
            student = db.session.get(User, student_id)
            self.assertFalse(student.monthly_fee_exempt)
            self.assertFalse(student.is_fee_exempt_for(current_date.year, current_date.month))
            self.assertEqual(student.get_numeric_price(current_date.year, current_date.month), 100.0)
            self.assertEqual(student.get_month_schedule(current_date.month, current_date.year)[current_date.month - 1]['status'], 'atrasado')

    def test_sidebar_only_renders_items_allowed_for_each_role(self):
        self.login('aluno')
        student_menu = self.client.get('/dashboard').get_data(as_text=True)
        self.assertIn('Calendário de Aulas', student_menu)
        self.assertNotIn('CAMPEONATOS', student_menu)
        self.assertNotIn('Loja Oficial', student_menu)
        self.assertNotIn('Pesos', student_menu)
        self.assertNotIn('Placar e Timer', student_menu)
        self.assertNotIn('Dashboard Financeiro', student_menu)
        self.assertNotIn('Cards Planos', student_menu)
        self.assertNotIn('Gestão de Planos', student_menu)
        self.assertNotIn('Gestão de Mensalidades', student_menu)
        self.assertNotIn('Gestão de Privilégios', student_menu)
        self.assertNotIn('CENTRAL DE INTEGRAÇÕES', student_menu)

        with self.client.session_transaction() as data:
            data.clear()
        self.login('monitor')
        monitor_menu = self.client.get('/dashboard').get_data(as_text=True)
        self.assertNotIn('Confirmar Presenças', monitor_menu)
        self.assertIn('Presenças & Treinos', monitor_menu)
        self.assertIn('Combos dos Alunos', monitor_menu)
        self.assertNotIn('Dashboard Financeiro', monitor_menu)
        self.assertNotIn('Gestão de Mensalidades', monitor_menu)
        self.assertNotIn('Gestão de Planos', monitor_menu)
        self.assertNotIn('Gestão de Turmas', monitor_menu)
        self.assertNotIn('Gestão de Privilégios', monitor_menu)
        self.assertNotIn('Administração de Graduação', monitor_menu)
        self.assertNotIn('CAMPEONATOS', monitor_menu)
        self.assertNotIn('Loja Oficial', monitor_menu)
        self.assertNotIn('Pesos', monitor_menu)
        self.assertNotIn('Placar e Timer', monitor_menu)
        self.assertNotIn('Integrações • Pagamentos', monitor_menu)
        self.assertNotIn('Integrações • Catraca', monitor_menu)
        self.assertNotIn('/mensalidades_admin.html', monitor_menu)

        with self.client.session_transaction() as data:
            data.clear()
        self.login('instrutor')
        instructor_menu = self.client.get('/dashboard').get_data(as_text=True)
        self.assertIn('FINANCEIRO E ADMINISTRAÇÃO', instructor_menu)
        self.assertIn('Visão Financeira', instructor_menu)
        self.assertIn('Gestão de Turmas', instructor_menu)
        self.assertIn('Planos', instructor_menu)
        self.assertIn('Mensalidades', instructor_menu)
        self.assertIn('Usuários e Permissões', instructor_menu)
        self.assertIn('Graduações', instructor_menu)
        self.assertNotIn('Cards Planos', instructor_menu)
        self.assertNotIn('Confirmar Presenças', instructor_menu)
        self.assertNotIn('CAMPEONATOS', instructor_menu)
        self.assertNotIn('Loja Oficial', instructor_menu)
        self.assertNotIn('Interno', instructor_menu)
        self.assertNotIn('Placar e Timer', instructor_menu)
        self.assertIn('INTEGRAÇÕES', instructor_menu)
        self.assertIn('Catraca', instructor_menu)
        self.assertIn('/mensalidades_admin.html', instructor_menu)
        payments_page = self.client.get('/integracoes/pagamentos').get_data(as_text=True)
        self.assertIn('CENTRAL DE INTEGRAÇÕES', payments_page)
        self.assertIn('Fluxo automático da mensalidade', payments_page)
        self.assertIn('Como faremos a implantação', payments_page)
        self.assertIn('O dinheiro passa pelo servidor do BJ Sports?', payments_page)
        self.assertIn('nenhum provedor está conectado', payments_page)
        turnstile_page = self.client.get('/integracoes/catraca').get_data(as_text=True)
        self.assertIn('Integrações • Catraca', turnstile_page)
        self.assertIn('FORMAS DE IDENTIFICAÇÃO', turnstile_page)
        self.assertIn('REGRAS DE LIBERAÇÃO', turnstile_page)
        self.assertIn('Topdata Fit Easy com leitor facial', turnstile_page)
        self.assertIn('APIs abertas informadas', turnstile_page)
        self.assertIn('Nenhuma catraca está sendo comandada', turnstile_page)
        championship_page = self.client.get('/campeonatos/interno').get_data(as_text=True)
        self.assertIn('Campeonatos internos', championship_page)
        self.assertIn('Novo campeonato', championship_page)
        self.assertIn('Nenhum campeonato publicado', championship_page)

        classes_page = self.client.get('/gestao/turmas').get_data(as_text=True)
        self.assertIn('GESTÃO OPERACIONAL', classes_page)
        self.assertIn('Turmas e capacidade', classes_page)
        self.assertIn('Dados insuficientes', classes_page)
        self.assertIn('Sem dados', classes_page)
        self.assertNotIn('Vínculos iniciais de demonstração', classes_page)
        self.assertIn('data-class-modal', classes_page)
        filtered_classes = self.client.get('/gestao/turmas?modality=Boxe').get_data(as_text=True)
        self.assertIn('Boxe Matinal', filtered_classes)
        self.assertNotIn('Jiu-Jitsu Kids 1', filtered_classes)

    def test_instructor_can_save_class_changes_and_open_management_details(self):
        self.login('instrutor')
        self.client.get('/gestao/turmas')
        with app.app_context():
            class_group = ClassGroup.query.order_by(ClassGroup.id).first()
            student = User.query.filter_by(role='aluno').first()
            db.session.add(ClassEnrollment(user_id=student.id, class_group_id=class_group.id, active=True))
            db.session.commit()
            class_id = class_group.id
            monitor = User.query.filter_by(role='monitor').first()
            monitor_id = monitor.id
            monitor_name = monitor.name

        response = self.client.post('/gestao/turmas', data={
            'action': 'update', 'class_id': class_id, 'class_name': 'Turma Atualizada',
            'class_modality': class_group.modality, 'class_audience': 'Adulto',
            'class_schedule': 'Seg, Qua, Sex • 06:30', 'class_instructor': 'Mestre Bolivar',
            'class_capacity': '22', 'class_duration': '75', 'class_status': 'ativa',
            'responsible_monitor_id': str(monitor_id),
            'publish_public': '1', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('Turma Turma Atualizada salva com sucesso!', response.get_data(as_text=True))
        with app.app_context():
            class_group = db.session.get(ClassGroup, class_id)
            self.assertEqual(class_group.name, 'Turma Atualizada')
            self.assertEqual(class_group.capacity, 22)
            self.assertEqual(class_group.duration_minutes, 75)
            self.assertEqual(class_group.schedules, ['Seg, Qua, Sex • 06:30'])
            self.assertEqual(class_group.responsible_monitor_id, monitor_id)

        detail = self.client.get(f'/gestao/turmas/{class_id}')
        self.assertEqual(detail.status_code, 200)
        detail_page = detail.get_data(as_text=True)
        self.assertIn('Análise financeira, frequência e alunos vinculados', detail_page)
        self.assertIn('Mensalidades recebidas por competência', detail_page)
        self.assertIn('Taxa de presença', detail_page)
        self.assertNotIn('Vínculos iniciais de demonstração', detail_page)
        self.assertIn('Monitor responsável', detail_page)
        self.assertIn(monitor_name, detail_page)

    def test_monitor_can_be_responsible_for_multiple_classes(self):
        self.login('instrutor')
        self.client.get('/gestao/turmas')
        with app.app_context():
            monitor_id = User.query.filter_by(role='monitor').first().id
            classes = ClassGroup.query.order_by(ClassGroup.id).limit(2).all()
            payloads = [{
                'action': 'update', 'class_id': item.id, 'class_name': item.name,
                'class_modality': item.modality, 'class_audience': item.audience,
                'class_schedule': ' / '.join(item.schedules), 'class_instructor': item.instructor,
                'class_capacity': str(item.capacity), 'class_duration': str(item.duration_minutes),
                'class_status': item.status, 'responsible_monitor_id': str(monitor_id),
            } for item in classes]
            class_ids = [item.id for item in classes]

        for payload in payloads:
            payload['csrf_token'] = self.csrf()
            self.assertEqual(self.client.post('/gestao/turmas', data=payload).status_code, 302)

        with app.app_context():
            assigned = ClassGroup.query.filter_by(responsible_monitor_id=monitor_id).filter(ClassGroup.id.in_(class_ids)).count()
            self.assertEqual(assigned, 2)

    def test_csrf_is_required(self):
        self.assertEqual(self.client.post('/login', data={'action': 'login'}).status_code, 400)

    def test_store_routes_are_removed(self):
        self.assertEqual(self.client.get('/loja').status_code, 404)
        self.assertEqual(self.client.get('/loja.html').status_code, 404)

    def test_plan_admin_is_central_source_for_modalities_and_enrollments(self):
        self.login('instrutor')
        page = self.client.get('/planos_admin').get_data(as_text=True)
        self.assertIn('data-plan-tab="modalities"', page)
        self.assertIn('data-plan-tab="plans"', page)
        self.assertIn('class="plan-admin-table', page)
        self.assertNotIn('class="plan-admin-card"', page)
        self.assertNotIn('<th>Benefícios</th>', page)
        self.assertNotIn('<th>Nome</th>', page)
        self.assertNotIn('<th>Descrição</th>', page)
        self.assertIn('<thead><tr><th>Ações</th><th>Modalidades</th><th>Horários</th>', page)
        self.assertNotIn('plan-row-identity', page)
        self.assertIn('class="plan-modality-text"', page)
        self.assertIn('Planos totalmente personalizados', page)
        self.assertIn('<th>Plano personalizado</th><th>Modalidades incluídas</th>', page)
        self.assertNotIn('name="modalities" required aria-label="Modalidade"', page)
        self.assertIn('<th>Ter - Qui</th>', page)
        self.assertIn('<th>Seg - Qua - Sex</th>', page)
        self.assertIn('<th>Todos os dias</th>', page)
        self.assertIn('data-plan-benefits-toggle=', page)
        created = self.client.post('/planos_admin', data={
            'action': 'create', 'name': 'Plano Boxe Central', 'category': 'Planos Individuais',
            'price': 'R$ 110,00/mês', 'modalities': ['Boxe'], 'sub': 'Boxe oficial',
            'features': 'Técnica; Condicionamento', 'return_tab': 'modalities',
            'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('cadastrados. O catálogo já foi atualizado', created.get_data(as_text=True))
        self.assertEqual(created.request.args.get('tab'), 'modalities')
        with app.app_context():
            plan = Plan.query.filter_by(name='Plano Boxe Central').one()
            self.assertEqual(plan.modality, 'Boxe')
            student = User.query.filter_by(username='aluno').one()
            student.plan = 'Plano Boxe Central • Ter, Qui — R$ 110,00/mês'
            db.session.commit()
            plan_id = plan.id

        updated = self.client.post('/planos_admin', data={
            'action': 'update', 'plan_id': plan_id, 'name': 'Boxe Essencial',
            'category': 'Planos Individuais', 'price_ter_qui': 'R$ 105,00/mês',
            'price_seg_qua_sex': 'R$ 115,00/mês', 'price_all_days': 'R$ 130,00/mês',
            'modalities': ['Boxe'], 'sub': 'Plano atualizado', 'features': 'Técnica; Defesa',
            'return_tab': 'modalities', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('atualizado em todo o sistema e em 1 matrícula(s)', updated.get_data(as_text=True))
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            self.assertEqual(student.plan, 'Boxe Essencial • Ter, Qui — R$ 105,00/mês')
            plan = db.session.get(Plan, plan_id)
            self.assertEqual(plan.get_price_for_schedule('seg-qua-sex'), 'R$ 115,00/mês')
            self.assertEqual(plan.get_price_for_schedule('todos'), 'R$ 130,00/mês')

        blocked = self.client.post('/planos_admin', data={
            'action': 'delete', 'plan_id': plan_id, 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('Não é possível excluir', blocked.get_data(as_text=True))
        with app.app_context():
            self.assertIsNotNone(db.session.get(Plan, plan_id))

    def test_family_holder_links_dependent_by_cpf_prefix_without_individual_billing(self):
        with app.app_context():
            family = Plan(
                name='Plano Família', category='Combos & Planos Especiais', price='R$ 200,00/mês',
                price_all_days='R$ 200,00/mês', discount_percent=10, modality='Jiu-Jitsu, Boxe',
            )
            holder = User.query.filter_by(username='aluno').one()
            holder.plan = 'Plano Família • Todos os dias — R$ 200,00/mês'
            dependent = User(username='dependente', name='Dependente Teste', cpf='321.999.999-00',
                             ddd='83', phone='988888888', plan='Plano Teste — R$ 100,00/mês',
                             due_date='5', start_month=1, role='aluno', payment_status='Em Dia')
            dependent.set_password('senha-segura')
            db.session.add_all([family, dependent])
            db.session.commit()
            dependent_id = dependent.id

        self.login('aluno')
        response = self.client.post('/configuracoes', data={
            'action': 'add_plan_dependent', 'dependent_cpf3': '321', 'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('sem cobrança individual', response.get_data(as_text=True))
        with app.app_context():
            holder = User.query.filter_by(username='aluno').one()
            dependent = db.session.get(User, dependent_id)
            self.assertEqual(dependent.sponsor_id, holder.id)
            self.assertEqual(dependent.get_numeric_price(), 0)
            self.assertEqual(holder.get_numeric_price(), 180)

        self.client.post('/configuracoes', data={
            'action': 'remove_plan_dependent', 'dependent_id': dependent_id, 'csrf_token': self.csrf(),
        })
        with app.app_context():
            dependent = db.session.get(User, dependent_id)
            self.assertIsNone(dependent.sponsor_id)
            self.assertEqual(dependent.plan, 'Plano Teste — R$ 100,00/mês')

    def test_instructor_changes_student_plan_from_monthly_drawer_without_rewriting_history(self):
        self.login('instrutor')
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            combo = Plan(
                name='Plano Combo + 1', category='Combos & Planos Especiais',
                price='R$ 150,00/mês', modality='Jiu-Jitsu, Boxe, Muay Thai, MMA',
            )
            db.session.add(combo)
            db.session.flush()
            payment = MonthlyPayment(user_id=student.id, year=2026, month=1, status='pago', amount=100)
            db.session.add(payment)
            db.session.commit()
            student_id, combo_id, payment_id = student.id, combo.id, payment.id

        page = self.client.get('/mensalidades_admin').get_data(as_text=True)
        self.assertIn('Plano e modalidades do aluno', page)
        self.assertIn('Salvar novo plano', page)

        response = self.client.post('/mensalidades_admin', data={
            'action': 'update_student_plan', 'user_id': student_id, 'plan_id': combo_id,
            'training_days': 'ter-qui', 'selected_modalities': ['Jiu-Jitsu', 'Boxe'],
            'csrf_token': self.csrf(),
        }, follow_redirects=True)
        self.assertIn('Plano de Aluno alterado', response.get_data(as_text=True))
        with app.app_context():
            student = db.session.get(User, student_id)
            self.assertEqual(student.plan, 'Plano Combo + 1 • Todos os dias — R$ 150,00/mês')
            self.assertEqual(student.selected_modalities, 'Jiu-Jitsu, Boxe')
            self.assertEqual(float(db.session.get(MonthlyPayment, payment_id).amount), 100.0)

    def test_monthly_payment_has_year(self):
        self.login('instrutor')
        with app.app_context():
            student_id = User.query.filter_by(username='aluno').first().id
        response = self.client.post('/api/update_month_status', data={'user_id': student_id,
            'month': '8', 'year': '2027', 'status': 'pago', 'csrf_token': self.csrf()})
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            payment = MonthlyPayment.query.filter_by(user_id=student_id, year=2027, month=8).one()
            self.assertEqual(payment.status, 'pago')
            self.assertIsNotNone(payment.paid_at)

    def test_billing_message_keeps_approved_copy(self):
        with app.app_context():
            student = User.query.filter_by(username='aluno').first()
            student.payment_status = 'Pendente'
            student.set_month_status('8', 'atrasado', 2026)
            link = student.get_whatsapp_billing_link()
            self.assertIn('Segue%20o%20lembrete%20da%20sua%20mensalidade', link)
            self.assertIn('Chave%20PIX%20para%20pagamento', link)
            self.assertIn('Favor%20enviar%20o%20comprovante', link)

    def test_booking_is_persisted(self):
        response = self.client.post('/api/bookings', json={'login_or_name': 'Visitante Teste',
            'cpf3': '', 'modality': 'Jiu-Jitsu', 'shift_time': 'Segunda 19:00',
            'is_experimental': True}, headers={'X-CSRF-Token': self.csrf()})
        self.assertEqual(response.status_code, 201)
        with app.app_context():
            self.assertEqual(Booking.query.count(), 1)

    def test_student_with_debt_cannot_book_new_class(self):
        self.login('aluno')
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            student.payment_status = 'Pendente'
            db.session.commit()
        response = self.client.post('/api/bookings', json={
            'login_or_name': 'aluno', 'cpf3': '000', 'modality': 'Jiu-Jitsu',
            'shift_time': 'Segunda 19:00', 'is_experimental': False
        }, headers={'X-CSRF-Token': self.csrf()})
        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.get_json()['code'], 'payment_required')

    def test_student_with_debt_cannot_register_attendance(self):
        self.login('aluno')
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            student.payment_status = 'Pendente'
            db.session.commit()
        response = self.client.post('/presencas', data={'csrf_token': self.csrf()}, follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        page = response.get_data(as_text=True)
        self.assertIn('Check-in bloqueado', page)
        self.assertNotIn('Confirmar Check-in Hoje', page)
        with app.app_context():
            self.assertEqual(Attendance.query.count(), 0)
        with app.app_context():
            self.assertEqual(Booking.query.count(), 0)

    def test_overdue_student_sees_red_financial_indicator(self):
        with app.app_context():
            student = User.query.filter_by(username='aluno').first()
            student.payment_status = 'Pendente'
            db.session.commit()
        self.login('aluno')
        page = self.client.get('/mensalidades_aluno').get_data(as_text=True)
        self.assertIn('live-dot is-overdue', page)
        self.assertIn('EXISTEM MENSALIDADES EM ATRASO', page)
        self.assertIn('SITUAÇÃO FINANCEIRA: PENDENTE', page)

    def test_student_in_good_standing_can_book(self):
        self.login('aluno')
        response = self.client.post('/api/bookings', json={
            'login_or_name': 'aluno', 'cpf3': '000', 'modality': 'Jiu-Jitsu',
            'shift_time': 'Segunda 19:00', 'is_experimental': False
        }, headers={'X-CSRF-Token': self.csrf()})
        self.assertEqual(response.status_code, 201)

    def test_student_cannot_book_without_cpf_digits(self):
        response = self.client.post('/api/bookings', json={
            'login_or_name': 'aluno', 'modality': 'Jiu-Jitsu',
            'shift_time': 'Segunda 19:00', 'is_experimental': False
        }, headers={'X-CSRF-Token': self.csrf()})
        self.assertEqual(response.status_code, 400)

    def test_attendance_is_idempotent_per_day(self):
        self.login('aluno')
        for _ in range(2):
            self.assertEqual(self.client.post('/presencas', data={'csrf_token': self.csrf()}).status_code, 302)
        with app.app_context():
            self.assertEqual(Attendance.query.count(), 1)
            self.assertEqual(Attendance.query.one().status, 'pendente')

    def test_new_student_has_no_metrics_and_teacher_confirms_checkin(self):
        self.login('aluno')
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            student.plan = 'Jiu-Jitsu (Seg, Qua, Sex) — R$ 100,00/mês'
            db.session.commit()
        initial_page = self.client.get('/presencas').get_data(as_text=True)
        self.assertIn('Frequência em formação', initial_page)
        self.assertNotIn('Últimos 30 dias', initial_page)

        response = self.client.post('/presencas', data={
            'action': 'request_checkin', 'csrf_token': self.csrf()
        }, follow_redirects=True)
        self.assertIn('Aguardando confirmação', response.get_data(as_text=True))
        pending_dashboard = self.client.get('/dashboard').get_data(as_text=True)
        self.assertIn('0 neste mês', pending_dashboard)
        with app.app_context():
            attendance = Attendance.query.one()
            self.assertEqual(attendance.status, 'pendente')
            attendance_id = attendance.id

        with self.client.session_transaction() as data:
            data.clear()
        self.login('monitor')
        monitor_page = self.client.get('/presencas').get_data(as_text=True)
        self.assertIn('Check-ins pendentes', monitor_page)
        self.assertIn('Confirmar Presenças', monitor_page)
        self.assertIn('data-pending-attendance-count="1"', monitor_page)
        self.assertIn('erp-nav-pending-count">1', monitor_page)
        self.assertIn('attendance-approval-group', monitor_page)
        self.assertIn('Jiu-Jitsu', monitor_page)
        self.assertIn('Aluno', monitor_page)
        self.client.post('/presencas', data={
            'action': 'confirm_attendance', 'attendance_id': attendance_id,
            'csrf_token': self.csrf()
        })
        with app.app_context():
            attendance = db.session.get(Attendance, attendance_id)
            self.assertEqual(attendance.status, 'confirmado')
            self.assertEqual(attendance.confirmed_by_username, 'monitor')
            self.assertIsNotNone(attendance.confirmed_at)

    def test_instructor_can_reject_pending_checkin(self):
        self.login('aluno')
        self.client.post('/presencas', data={
            'action': 'request_checkin', 'csrf_token': self.csrf()
        })
        with app.app_context():
            attendance_id = Attendance.query.one().id

        with self.client.session_transaction() as data:
            data.clear()
        self.login('instrutor')
        teacher_page = self.client.get('/presencas').get_data(as_text=True)
        self.assertIn('Negar', teacher_page)
        self.assertIn('data-confirm-rejection', teacher_page)
        self.assertNotIn("confirm('Negar esta presença?')", teacher_page)
        response = self.client.post('/presencas', data={
            'action': 'reject_attendance', 'attendance_id': attendance_id,
            'csrf_token': self.csrf()
        }, follow_redirects=True)
        self.assertIn('Presença de Aluno negada.', response.get_data(as_text=True))
        with app.app_context():
            attendance = db.session.get(Attendance, attendance_id)
            self.assertEqual(attendance.status, 'negado')
            self.assertEqual(attendance.confirmed_by_username, 'instrutor')
            self.assertEqual(Attendance.query.filter_by(status='confirmado').count(), 0)

        with self.client.session_transaction() as data:
            data.clear()
        self.login('aluno')
        student_page = self.client.get('/presencas').get_data(as_text=True)
        self.assertIn('Check-in não confirmado', student_page)
        self.assertIn('não será contabilizado', student_page)

    def test_attendance_page_uses_real_windows_without_class_descriptions(self):
        self.login('aluno')
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            student.plan = 'Jiu-Jitsu (Seg, Qua, Sex) — R$ 100,00/mês'
            today = datetime.now().date()
            for days_ago in (5, 20, 40, 80, 120, 179):
                db.session.add(Attendance(user_id=student.id, training_date=today - timedelta(days=days_ago), status='confirmado'))
            db.session.commit()
        page = self.client.get('/presencas').get_data(as_text=True)
        self.assertIn('Últimos 30 dias', page)
        self.assertIn('Últimos 90 dias', page)
        self.assertIn('Últimos 180 dias', page)
        self.assertIn('2 de ', page)
        self.assertIn('4 de ', page)
        self.assertIn('6 de ', page)
        self.assertGreaterEqual(page.count('%</strong>'), 3)
        self.assertGreaterEqual(page.count('is-below-target'), 3)
        self.assertNotIn('Passagem de Guarda', page)
        self.assertNotIn('Armlock', page)
        self.assertNotIn('Sparring', page)

    def test_attendance_indicator_is_green_at_or_above_75_percent(self):
        self.login('aluno')
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            student.plan = 'Jiu-Jitsu (Seg, Qua, Sex) — R$ 100,00/mês'
            today = datetime.now().date()
            scheduled_dates = [
                today - timedelta(days=offset) for offset in range(30)
                if (today - timedelta(days=offset)).weekday() in {0, 2, 4}
            ]
            minimum_for_green = (len(scheduled_dates) * 3 + 3) // 4
            db.session.add(Attendance(user_id=student.id, training_date=today - timedelta(days=30), status='confirmado'))
            for training_date in scheduled_dates[:minimum_for_green]:
                db.session.add(Attendance(user_id=student.id, training_date=training_date, status='confirmado'))
            db.session.commit()
        page = self.client.get('/presencas').get_data(as_text=True)
        self.assertIn('attendance-summary-card is-30 is-good', page)

    def test_user_can_update_belt_color(self):
        self.login('aluno')
        response = self.client.post('/configuracoes', data={
            'name': 'Aluno Teste', 'phone': '(83) 999999999',
            'belt_color': 'roxa', 'belt_degree': '2', 'csrf_token': self.csrf()
        })
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            self.assertEqual((student.belt_color, student.belt_degree), ('roxa', 2))

    def test_contract_page_is_printable_and_records_updated_acceptance(self):
        self.login('aluno')
        page = self.client.get('/minha-conta/contrato.html').get_data(as_text=True)
        self.assertIn('Minha Conta • Contrato', page)
        self.assertIn('Atualização aguardando seu aceite', page)
        self.assertIn('data-contract-print', page)
        self.assertIn('Mudanças neste termo', page)
        self.assertIn('Aceitar atualização', page)
        self.assertIn('Autorização de imagem necessária', page)
        self.assertIn('name="imageConsentScope"', page)
        self.assertIn('data-contract-pending="true"', page)
        self.assertIn('Atualização aguardando aceite', page)

        denied = self.client.post('/minha-conta/contrato.html', data={
            'action': 'accept_contract_update', 'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('marque a confirmação', denied)

        denied_image = self.client.post('/minha-conta/contrato.html', data={
            'action': 'accept_contract_update', 'acceptContractUpdate': 'on',
            'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Autorize o uso de imagem', denied_image)

        accepted = self.client.post('/minha-conta/contrato.html', data={
            'action': 'accept_contract_update', 'acceptContractUpdate': 'on',
            'imageConsentScope': 'adult',
            'csrf_token': self.csrf(),
        }, follow_redirects=True).get_data(as_text=True)
        self.assertIn('Contrato atualizado', accepted)
        self.assertIn('data-contract-pending="false"', accepted)
        self.assertNotIn('Aceitar atualização</button>', accepted)
        with app.app_context():
            student = User.query.filter_by(username='aluno').one()
            self.assertEqual(student.membership_terms_version, MEMBERSHIP_TERMS_VERSION)
            self.assertEqual(student.privacy_notice_version, PRIVACY_NOTICE_VERSION)
            self.assertTrue(student.image_use_consent)
            self.assertEqual(student.image_consent_scope, 'adult')
            acceptance = ContractAcceptance.query.filter_by(user_id=student.id).one()
            self.assertEqual(acceptance.source, 'account_update')

    def test_cards_planos_page_accessible_to_logged_users(self):
        # Redirects if unauthenticated
        self.assertEqual(self.client.get('/cards_planos').status_code, 302)

        # Logged in student can view plan cards
        self.login('aluno')
        res = self.client.get('/cards_planos')
        self.assertEqual(res.status_code, 200)
        page = res.get_data(as_text=True)
        self.assertIn('Cards Planos', page)
        self.assertIn('Total de Planos', page)
        self.assertIn('Receita Estimada', page)

    def test_60h_free_trial_and_monitor_baixa_authorization(self):
        with app.app_context():
            # Create a new student (trial active)
            new_student = User(username='novato', name='Aluno Novato', cpf='111.222.333-44',
                               ddd='83', phone='988887777', plan='Jiu-Jitsu (Seg, Qua, Sex)',
                               due_date='5', start_month=1, role='aluno', payment_status='Pendente',
                               created_at=datetime.utcnow())
            new_student.set_password('senha-segura')
            db.session.add(new_student)
            db.session.commit()

            trial = new_student.get_trial_status()
            self.assertTrue(trial['in_trial'])
            self.assertFalse(trial['expired'])
            self.assertGreaterEqual(trial['hours_left'], 59)

            # Fast forward created_at to 70 hours ago
            new_student.created_at = datetime.utcnow() - timedelta(hours=70)
            db.session.commit()

            trial_expired = new_student.get_trial_status()
            self.assertFalse(trial_expired['in_trial'])
            self.assertTrue(trial_expired['expired'])

            # Test permission: Professor (instrutor) can manage any student
            prof = User.query.filter_by(role='instrutor').first()
            self.assertTrue(new_student.can_be_managed_by(prof))

            # Monitor without shared turma cannot manage student
            mon = User.query.filter_by(role='monitor').first()
            self.assertFalse(new_student.can_be_managed_by(mon))

            # Assign monitor to class group and enroll student
            cg = ClassGroup(name='Turma Jiu-Jitsu Monitor', modality='Jiu-Jitsu', audience='Adulto',
                            instructor=mon.name, capacity=20, duration_minutes=60)
            db.session.add(cg)
            db.session.commit()

            db.session.add(ClassEnrollment(user_id=new_student.id, class_group_id=cg.id, active=True))
            db.session.add(ClassEnrollment(user_id=mon.id, class_group_id=cg.id, active=True))
            db.session.commit()

            # Now monitor can manage student because they share the class group
            self.assertTrue(new_student.can_be_managed_by(mon))

    def test_dar_baixa_mensalidade_route(self):
        with app.app_context():
            student = User.query.filter_by(username='aluno').first()
            student.payment_status = 'Pendente'
            db.session.commit()

            # Professor logs in and confirms payment
            self.login('instrutor')
            res = self.client.post(f'/dar-baixa-mensalidade/{student.id}', data={'csrf_token': self.csrf()}, follow_redirects=True)
            self.assertEqual(res.status_code, 200)

            updated_student = db.session.get(User, student.id)
            self.assertEqual(updated_student.payment_status, 'Em Dia')


if __name__ == '__main__':
    unittest.main()
